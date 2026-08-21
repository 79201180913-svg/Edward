from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

HEADERS = [
    'date', 'time', 'account_id', 'order_id', 'instrument_uid', 'FIGI', 'ticker', 'name',
    'operation', 'quantity', 'order_type', 'execution_price', 'amount', 'commission',
    'currency', 'status'
]


@dataclass(frozen=True, slots=True)
class TradeRecord:
    account_id: str
    order_id: str
    instrument_uid: str
    operation: str
    quantity: int
    order_type: str
    execution_price: Decimal | None
    amount: Decimal | None
    commission: Decimal | None
    currency: str
    status: str
    figi: str = ''
    ticker: str = ''
    name: str = ''
    executed_at: datetime | None = None


class TradingHistoryRepository:
    def __init__(self, path: str | Path = 'data/trading_history.xlsx') -> None:
        self.path = Path(path)

    def save_completed(self, record: TradeRecord) -> None:
        if record.status != 'FILLED':
            raise ValueError('save_completed accepts only FILLED operations')
        self.save(record)

    def save(self, record: TradeRecord) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        workbook, sheet = self._open_sheet()
        try:
            if self._contains_order(sheet, record.order_id):
                return
            self._append(sheet, record)
            workbook.save(self.path)
        finally:
            workbook.close()

    def upsert(self, record: TradeRecord) -> None:
        """Insert a new order or update the existing row with the same order_id."""
        self.path.parent.mkdir(parents=True, exist_ok=True)
        workbook, sheet = self._open_sheet()
        try:
            order_column = HEADERS.index('order_id') + 1
            target_row = None
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=order_column).value == record.order_id:
                    target_row = row
                    break

            timestamp = record.executed_at or datetime.now(timezone.utc)
            values = {
                'date': timestamp.date().isoformat(),
                'time': timestamp.time().isoformat(timespec='seconds'),
                'account_id': record.account_id,
                'order_id': record.order_id,
                'instrument_uid': record.instrument_uid,
                'FIGI': record.figi,
                'ticker': record.ticker,
                'name': record.name,
                'operation': record.operation,
                'quantity': record.quantity,
                'order_type': record.order_type,
                'execution_price': self._decimal(record.execution_price),
                'amount': self._decimal(record.amount),
                'commission': self._decimal(record.commission),
                'currency': record.currency,
                'status': record.status,
            }
            row_values = [values[h] for h in HEADERS]
            if target_row is None:
                sheet.append(row_values)
            else:
                for column, value in enumerate(row_values, start=1):
                    sheet.cell(row=target_row, column=column).value = value
            workbook.save(self.path)
        finally:
            workbook.close()

    def _open_sheet(self) -> tuple[Any, Any]:
        if self.path.exists():
            workbook = load_workbook(self.path)
            sheet = workbook.active
            self._ensure_headers(sheet)
            return workbook, sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Trades'
        sheet.append(HEADERS)
        return workbook, sheet

    def _append(self, sheet: Any, record: TradeRecord) -> None:
        timestamp = record.executed_at or datetime.now(timezone.utc)
        data = {
            'date': timestamp.date().isoformat(),
            'time': timestamp.time().isoformat(timespec='seconds'),
            'account_id': record.account_id,
            'order_id': record.order_id,
            'instrument_uid': record.instrument_uid,
            'FIGI': record.figi,
            'ticker': record.ticker,
            'name': record.name,
            'operation': record.operation,
            'quantity': record.quantity,
            'order_type': record.order_type,
            'execution_price': self._decimal(record.execution_price),
            'amount': self._decimal(record.amount),
            'commission': self._decimal(record.commission),
            'currency': record.currency,
            'status': record.status,
        }
        sheet.append([data[h] for h in HEADERS])

    def read_all(self) -> list[dict[str, Any]]:
        if not self.path.exists():
            return []
        workbook = load_workbook(self.path, read_only=True, data_only=True)
        try:
            rows = list(workbook.active.iter_rows(values_only=True))
        finally:
            workbook.close()
        return [dict(zip(HEADERS, row)) for row in rows[1:]] if rows else []

    @staticmethod
    def _decimal(value: Decimal | None) -> float | None:
        return float(value) if value is not None else None

    @staticmethod
    def _ensure_headers(sheet: Any) -> None:
        existing = [cell.value for cell in sheet[1]]
        if existing != HEADERS:
            for index, header in enumerate(HEADERS, start=1):
                sheet.cell(row=1, column=index).value = header

    @staticmethod
    def _contains_order(sheet: Any, order_id: str) -> bool:
        column = HEADERS.index('order_id') + 1
        return any(
            cell.value == order_id
            for cell in next(sheet.iter_cols(min_col=column, max_col=column, min_row=2, values_only=False))
        )
