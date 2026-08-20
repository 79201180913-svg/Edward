from decimal import Decimal

from edward.domain.order_state import OrderSnapshot, OrderStatus
from edward.history.trading_history import TradeRecord, TradingHistoryRepository
from edward.services.execution_service import ExecutionContext, ExecutionService


def test_filled_trade_is_saved_once(tmp_path):
    path = tmp_path / "history.xlsx"
    repository = TradingHistoryRepository(path)
    service = ExecutionService(repository)
    snapshot = OrderSnapshot(
        order_id="order-1",
        account_id="acc-1",
        instrument_uid="uid-1",
        status=OrderStatus.FILLED,
        requested_quantity=10,
        filled_quantity=10,
        remaining_quantity=0,
        average_fill_price=Decimal("100.50"),
        commission=Decimal("1.20"),
    )
    context = ExecutionContext(operation="BUY", order_type="MARKET", currency="RUB")
    service.process(snapshot, context)
    service.process(snapshot, context)

    from openpyxl import load_workbook
    workbook = load_workbook(path)
    sheet = workbook.active
    assert sheet.max_row == 2
    assert sheet.cell(2, 4).value == "order-1"
    assert sheet.cell(2, 12).value == 100.5
    assert sheet.cell(2, 14).value == 1.2
    workbook.close()
